"""Read / summarise Excel ``.xlsx`` workbooks for Jarvis -- structured-data handling.

``read_csv`` (Phase 30) reads plain-text CSV/TSV data files, but it deliberately
REFUSES a binary Excel workbook and tells the user to "save it as CSV" -- an
annoying dead-end, because a spreadsheet is exactly the kind of file a real user
keeps in Excel (budgets, expenses, contact lists, exports from other apps). This
closes that gap: ``read_excel`` opens a ``.xlsx`` workbook and measures it
EXACTLY -- how many sheets it has, and for a chosen sheet how many data rows and
columns, the column names, and a preview of the first rows ("how many rows are
in my workbook.xlsx", "what columns are in my budget", "read sheet 2",
"summarise my expenses spreadsheet"). The natural partner to ``find_files``
(locate the workbook, then read it) and the Excel counterpart to ``read_csv``.

Dependency: this needs ``openpyxl`` -- a well-established, PURE-PYTHON, offline
package (no compiler, no network). It is imported LAZILY inside the tool so it
never slows Jarvis's startup, and if it is somehow missing at runtime the tool
degrades to a friendly "install openpyxl" message rather than crashing. Jarvis
stays fully local/offline.

Safety model (strict, because an 8B local model WILL eventually pass junk, the
wrong type, or point this at something enormous):

- **Rooted in the user's home only.** The path is resolved and REJECTED unless
  it lives inside the user's home directory (same boundary as ``find_files``),
  so the model can never read a workbook from ``C:\\Windows``.
- **Bounded everywhere.** The file on disk is capped before opening, the workbook
  is opened in openpyxl's streaming READ-ONLY mode, the row scan is capped (a
  huge sheet stops early with a note), and every listed column name / preview
  cell is truncated -- a giant or hostile file can't exhaust memory or flood the
  agent's context.
- **Pure ASCII out.** Sheet names, column names and cell values are forced to
  single-line ASCII, so the summary can never corrupt the console/context.
- **Never raises.** A corrupt/non-xlsx file, an old ``.xls`` binary, a
  password-protected workbook, a missing dependency, a missing sheet, a missing
  file, a folder, an empty workbook, the wrong type, an empty or missing path --
  every one comes back as a friendly, pure-ASCII string the model can recover
  from.
"""

import datetime
import warnings
from pathlib import Path

from .document import _ascii_body
from .find import _coerce
from .organize import _ascii, _first_str, _resolve_under_home
from .registry import tool

MAX_PATH_LEN = 400                    # a path, not an essay
MAX_FILE_BYTES = 25 * 1024 * 1024     # refuse a workbook bigger than this on disk
MAX_ROWS = 200_000                    # most rows scanned before stopping (with a note)
DEFAULT_PREVIEW = 5                   # preview rows shown by default
MAX_PREVIEW_ROWS = 20                 # most preview rows the model may ask for
MAX_COLS_LISTED = 40                  # most column names listed
MAX_CELL = 40                         # truncate any single cell value to this
MAX_COLS_SCAN = 1000                  # ignore absurd column counts in a row
MAX_SHEETS_LISTED = 40                # most sheet names listed


def _import_openpyxl():
    """Import openpyxl lazily. Returns the module, or None if it isn't installed.
    Kept as a tiny seam so startup pays nothing and the smoke test can simulate
    a missing dependency."""
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except Exception:
        return None


def _flat(text: str) -> str:
    """Transliterate to readable single-line ASCII: curly quotes/dashes become
    plain ASCII and accents are stripped (cafe, not caf?) via read_document's
    transliterator, then any newline/tab is flattened to a space. Unlike the
    plain ``_ascii`` (which turns every accent into '?'), real spreadsheet text
    with accented names or smart quotes stays readable."""
    body = _ascii_body(text)  # transliterate; keeps newlines/tabs
    return " ".join(body.split())


def _cell(value) -> str:
    """One spreadsheet cell value, rendered to safe single-line ASCII + truncated.
    Handles the types openpyxl hands back (str/int/float/bool/datetime/None)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, float):
        # show an integral float (1200.0) as a plain int; keep real decimals
        text = str(int(value)) if value.is_integer() else repr(value)
    elif isinstance(value, datetime.datetime):
        # drop a midnight time so a pure date reads as a date, not "... 00:00:00"
        if value.hour == value.minute == value.second == value.microsecond == 0:
            text = value.date().isoformat()
        else:
            text = value.isoformat(sep=" ")
    elif isinstance(value, datetime.date):
        text = value.isoformat()
    else:
        text = str(value)
    v = _flat(text)
    if len(v) > MAX_CELL:
        v = v[:MAX_CELL - 3] + "..."
    return v


def _preview_int(value) -> int:
    """Coerce the optional 'rows' preview count into 0..MAX_PREVIEW_ROWS."""
    s = _coerce(value, 12)
    if not s:
        return DEFAULT_PREVIEW
    digits = "".join(ch for ch in s if ch.isdigit())  # tolerate "5 rows"
    if not digits:
        return DEFAULT_PREVIEW
    try:
        n = int(digits)
    except ValueError:
        return DEFAULT_PREVIEW
    if n < 0:
        return DEFAULT_PREVIEW
    return min(n, MAX_PREVIEW_ROWS)


def _pick_sheet(wb, want):
    """Choose a worksheet. ``want`` may be empty (active/first sheet), a sheet
    NAME, or a 1-based sheet NUMBER. Returns (worksheet, "") or (None, error)."""
    names = list(wb.sheetnames)
    if not names:
        return None, "the workbook has no sheets, sir."
    sel = _coerce(want, 200).strip() if want is not None else ""
    if not sel:
        return wb[names[0]], ""
    # exact name match first (case-sensitive, then case-insensitive)
    if sel in names:
        return wb[sel], ""
    low = {n.lower(): n for n in names}
    if sel.lower() in low:
        return wb[low[sel.lower()]], ""
    # a bare number picks the Nth sheet (1-based)
    if sel.isdigit():
        idx = int(sel)
        if 1 <= idx <= len(names):
            return wb[names[idx - 1]], ""
        return None, (f"there's no sheet {idx}, sir; the workbook has "
                      f"{len(names)} ({', '.join(_flat(n) for n in names[:MAX_SHEETS_LISTED])}).")
    shown = ", ".join(_flat(n) for n in names[:MAX_SHEETS_LISTED])
    return None, (f"there's no sheet called '{_ascii(sel)}', sir; the workbook "
                  f"has: {shown}.")


@tool(
    "read_excel",
    "Read and summarise an Excel .xlsx workbook: reports how many sheets it has "
    "and, for a chosen sheet, how many rows of data and columns, the column "
    "names, and a preview of the first rows. Use this whenever the user asks "
    "about a .xlsx file ('how many rows are in my workbook', 'what columns are "
    "in my budget', 'read sheet 2', 'summarise my expenses spreadsheet'); "
    "read_csv only handles plain-text CSV/TSV and your own counting is "
    "unreliable, this is exact. Give path (locate it first with find_files if "
    "you don't have it), optionally sheet (a sheet name or number; default the "
    "first sheet) and rows (how many preview rows to show). Only the user's own "
    "folders are allowed.",
    {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "The workbook to read, e.g. 'Documents/budget.xlsx'.",
            },
            "sheet": {
                "type": "string",
                "description": "Which sheet to read: a sheet name or number (default the first).",
            },
            "rows": {
                "type": "integer",
                "description": "How many preview rows to show (default 5).",
            },
        },
        "required": ["path"],
    },
)
def read_excel(path: str = "", sheet=None, rows=None, **extra) -> str:
    raw = _first_str(path, extra.get("file"), extra.get("document"),
                     extra.get("source"), extra.get("name"),
                     extra.get("filename"), extra.get("workbook"),
                     extra.get("spreadsheet"), extra.get("excel"))
    raw = _coerce(raw, MAX_PATH_LEN)
    if not raw:
        return "Error: tell me which Excel workbook to read, sir."

    want_sheet = sheet if sheet is not None else extra.get("tab")
    if want_sheet is None:
        want_sheet = extra.get("worksheet")
    want_rows = _preview_int(rows if rows is not None else extra.get("preview"))

    p, err = _resolve_under_home(raw)
    if p is None:
        return err or "Error: that file path isn't valid, sir."
    if not p.exists():
        return f"Error: I can't find '{_ascii(str(p))}', sir."
    if p.is_dir():
        return (f"Error: '{_ascii(p.name)}' is a folder, sir; give me an Excel "
                "workbook to read.")

    suffix = p.suffix.lower()
    if suffix in (".csv", ".tsv", ".tab"):
        return (f"Error: '{_ascii(p.name)}' is a text data file, sir; use "
                "read_csv for that. read_excel is for Excel .xlsx workbooks.")
    if suffix == ".xls":
        return (f"Error: '{_ascii(p.name)}' is an old Excel format I can't read, "
                "sir; open it in Excel and save it as .xlsx, then I'll read that.")
    if suffix not in (".xlsx", ".xlsm"):
        return (f"Error: '{_ascii(p.name)}' isn't an Excel workbook, sir; "
                "read_excel is for .xlsx files.")

    try:
        size = p.stat().st_size
    except OSError:
        size = 0
    if size > MAX_FILE_BYTES:
        return (f"Error: '{_ascii(p.name)}' is too large for me to read safely, "
                "sir.")

    openpyxl = _import_openpyxl()
    if openpyxl is None:
        return ("Error: I can't read Excel workbooks yet, sir -- the 'openpyxl' "
                "package isn't installed. Install it with: pip install openpyxl")

    wb = None
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")  # keep the console clean
            try:
                # read_only streams the sheet (bounded memory); data_only returns
                # cached values, not formula text
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            except Exception:
                return (f"Error: '{_ascii(p.name)}' isn't a valid Excel workbook, "
                        "sir; it may be corrupt, password-protected, or not "
                        "really an .xlsx.")

            sheet_names = list(wb.sheetnames)
            ws, serr = _pick_sheet(wb, want_sheet)
            if ws is None:
                return f"Error: {serr}"
            chosen = ws.title

            header: list[str] = []
            preview: list[list[str]] = []
            data_rows = 0
            scanned = 0
            capped = False
            for row in ws.iter_rows(values_only=True):
                scanned += 1
                if scanned > MAX_ROWS:
                    capped = True
                    break
                cells = list(row)[:MAX_COLS_SCAN]
                if not header:
                    # skip fully-blank leading rows so the header is the real one
                    if not any(c is not None and str(c).strip() for c in cells):
                        continue
                    header = [_cell(c) for c in cells]
                    continue
                # a fully blank row isn't counted so the total stays honest
                if not any(c is not None and str(c).strip() for c in cells):
                    continue
                data_rows += 1
                if len(preview) < want_rows:
                    preview.append([_cell(c) for c in cells])
    except Exception as e:
        return f"Error: couldn't read that workbook, sir ({_ascii(str(e))})."
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    def s(n):
        return "" if n == 1 else "s"

    n_sheets = len(sheet_names)
    sheet_list = ", ".join(_flat(n) for n in sheet_names[:MAX_SHEETS_LISTED])
    sheets_note = "" if n_sheets <= MAX_SHEETS_LISTED else \
        f" (+{n_sheets - MAX_SHEETS_LISTED} more)"
    book_line = (f"{_ascii(p.name)} -- {n_sheets} sheet{s(n_sheets)} "
                 f"({sheet_list}{sheets_note}).")

    if not header:
        return (book_line + f" Sheet '{_ascii(chosen)}' is empty, sir; there's "
                "nothing to read.")

    ncols = len(header)
    shown_cols = header[:MAX_COLS_LISTED]
    cols_note = "" if ncols <= MAX_COLS_LISTED else f" (+{ncols - MAX_COLS_LISTED} more)"
    row_note = " (stopped early; more rows remain)" if capped else ""

    head = (f"{book_line} Sheet '{_ascii(chosen)}': {data_rows} data "
            f"row{s(data_rows)}, {ncols} column{s(ncols)}.{row_note} "
            f"Columns: {', '.join(shown_cols)}{cols_note}.")

    if not preview:
        return head + " No data rows to preview."

    lines = [head, f"First {len(preview)} row{s(len(preview))}:"]
    for i, row in enumerate(preview, 1):
        lines.append(f"{i}: " + " | ".join(row))
    return "\n".join(lines)
